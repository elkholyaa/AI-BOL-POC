# excel_utils.py
"""
Purpose:
  This file provides utility functions for Excel file manipulation within the AI-BOL-POC application.
  Specifically, it contains a function to append a row to an existing Excel file ("bol.xlsx") while attempting
  to preserve the cell formatting of the previous rows.

Role:
  - Supports the proof-of-concept by enabling the addition of new rows (e.g., Bill of Lading data) to an Excel file.
  - Can be imported into other parts of the application where Excel updates are needed.

Workflow:
  - The function attempts to load an existing Excel file. If it doesn't exist, a new workbook is created and a header row is added.
  - The new row is appended in the order of the header columns.
  - If there is at least one data row present, the function copies the style from the previous row into the newly appended row.
  - Finally, the workbook is saved.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

def append_row_to_excel(row_data, file_path="bol.xlsx"):
    """
    Appends a new row of data to the specified Excel file while preserving cell formatting.

    Parameters:
      row_data (dict): A dictionary where keys are column names and values are the data to append.
      file_path (str): The path to the Excel file to update. Defaults to "bol.xlsx".

    Workflow:
      - Loads the existing Excel file or creates a new workbook if the file doesn't exist.
      - If creating a new file, writes a header row based on the keys of row_data.
      - Orders the new row data according to the header row.
      - Appends the new row.
      - Copies the cell formatting from the previous row to the new row (if applicable).
      - Saves the updated workbook.
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
    except (FileNotFoundError, InvalidFileException):
        # Create a new workbook and add header row if file does not exist or is invalid.
        wb = Workbook()
        ws = wb.active
        headers = list(row_data.keys())
        ws.append(headers)

    # Retrieve the header row from the first row.
    header_row = [cell.value for cell in ws[1]]
    # Order the new row data according to the header.
    new_row = [row_data.get(col, "") for col in header_row]

    # Get current last row index before appending.
    previous_last_row_idx = ws.max_row
    ws.append(new_row)
    new_row_idx = ws.max_row  # Index of the newly appended row

    # If there is a previous data row, copy its style to the new row.
    if previous_last_row_idx >= 2:
        for col_idx, new_cell in enumerate(ws[new_row_idx], start=1):
            prev_cell = ws.cell(row=previous_last_row_idx, column=col_idx)
            if prev_cell.has_style:
                new_cell.font = prev_cell.font
                new_cell.border = prev_cell.border
                new_cell.fill = prev_cell.fill
                new_cell.number_format = prev_cell.number_format
                new_cell.protection = prev_cell.protection
                new_cell.alignment = prev_cell.alignment

    wb.save(file_path)

# Example usage for testing the function.
if __name__ == "__main__":
    sample_row = {
        "Bill of Lading No.": "MSCUKK871206",
        "Shipper": "ABC Shipping",
        "Consignee": "XYZ Importers",
        "Port of Loading": "Port A",
        "Port of Discharge": "Port B"
    }
    append_row_to_excel(sample_row)
